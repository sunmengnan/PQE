#!/usr/bin/env python3
"""PQE Phase 1 MVP: parse CPK/FAI Excel reports, calculate metrics, export summary workbook.

This script intentionally reads Excel files through their OOXML XML parts instead of relying on
Excel/LibreOffice formula recalculation. It recalculates the key statistics from raw samples so
the output is deterministic and template-cache independent.
"""

import argparse
import json
import math
import re
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def col_to_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch.upper()) - 64
    return n


def as_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in ("", " "):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_parenthesized_parts(filename: str) -> List[str]:
    return [part.strip() for part in re.findall(r"\(([^()]*)\)", filename) if part.strip()]


def file_tag_text(filename: str) -> str:
    return " | ".join(extract_parenthesized_parts(filename))


def norm_cdf(z: float) -> float:
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


def safe_stdev(values: List[float]) -> Optional[float]:
    if len(values) < 2:
        return None
    try:
        return statistics.stdev(values)
    except statistics.StatisticsError:
        return None


class ExcelXmlReader:
    def __init__(self, path: Path):
        self.path = path
        self.zip = ZipFile(path)
        self.shared_strings = self._read_shared_strings()
        self.sheet_targets = self._read_sheet_targets()

    def close(self) -> None:
        self.zip.close()

    def __enter__(self) -> "ExcelXmlReader":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def _read_shared_strings(self) -> List[str]:
        if "xl/sharedStrings.xml" not in self.zip.namelist():
            return []
        root = ET.fromstring(self.zip.read("xl/sharedStrings.xml"))
        strings = []
        for si in root.findall(NS_MAIN + "si"):
            strings.append("".join(t.text or "" for t in si.iter(NS_MAIN + "t")))
        return strings

    def _read_sheet_targets(self) -> Dict[str, str]:
        workbook = ET.fromstring(self.zip.read("xl/workbook.xml"))
        rels = ET.fromstring(self.zip.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        targets = {}
        for sheet in workbook.find(NS_MAIN + "sheets"):
            name = sheet.attrib["name"]
            rid = sheet.attrib[NS_REL + "id"]
            target = rid_to_target[rid]
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")
            targets[name] = target
        return targets

    def sheet_names(self) -> List[str]:
        return list(self.sheet_targets.keys())

    def cell_value(self, cell: ET.Element) -> Any:
        cell_type = cell.attrib.get("t")
        if cell_type == "inlineStr":
            return "".join(t.text or "" for t in cell.iter(NS_MAIN + "t"))
        v = cell.find(NS_MAIN + "v")
        if v is None:
            return ""
        text = v.text or ""
        if cell_type == "s":
            try:
                return self.shared_strings[int(text)]
            except (ValueError, IndexError):
                return text
        if cell_type == "b":
            return text == "1"
        try:
            return float(text) if any(ch in text for ch in ".Ee") else int(text)
        except ValueError:
            return text

    def iter_rows(self, sheet_name: str) -> Iterable[Tuple[int, Dict[int, Any]]]:
        target = self.sheet_targets[sheet_name]
        root = ET.fromstring(self.zip.read(target))
        sheet_data = root.find(NS_MAIN + "sheetData")
        if sheet_data is None:
            return
        for row in sheet_data.findall(NS_MAIN + "row"):
            row_num = int(row.attrib.get("r", "0"))
            values = {}
            for cell in row.findall(NS_MAIN + "c"):
                ref = cell.attrib.get("r", "")
                match = re.match(r"([A-Z]+)", ref)
                if not match:
                    continue
                values[col_to_num(match.group(1))] = self.cell_value(cell)
            yield row_num, values


def normalize_cavity(sheet_name: str) -> str:
    cav_match = re.search(r"CAV\s*(\d+)", sheet_name, flags=re.I)
    if cav_match:
        return "CAV%s" % int(cav_match.group(1))
    n_match = re.match(r"\s*(N[A-Za-z])\s*(\d+)", sheet_name, flags=re.I)
    if n_match:
        return "%s%s" % (n_match.group(1), int(n_match.group(2)))
    return sheet_name.strip()


def compute_limits(dim_type: str, nominal: Optional[float], tol_plus: Optional[float], tol_minus: Optional[float]) -> Tuple[Optional[float], Optional[float]]:
    dim = dim_type.strip().upper()
    if dim in ("GD&T", "GDT"):
        # GD&T values in these templates are non-negative deviation values.
        return None, tol_plus
    if dim == "MIN":
        lower = tol_plus if tol_plus is not None else nominal
        return lower, None
    if dim == "MAX":
        upper = tol_plus if tol_plus is not None else nominal
        return None, upper
    if nominal is None:
        return None, None
    lower = nominal + tol_minus if tol_minus is not None else None
    upper = nominal + tol_plus if tol_plus is not None else None
    return lower, upper


def calc_stats(samples: List[float], lower: Optional[float], upper: Optional[float]) -> Dict[str, Any]:
    stats: Dict[str, Any] = {
        "sample_count": len(samples),
        "max": None,
        "min": None,
        "mean": None,
        "median": None,
        "stddev": None,
        "cp": None,
        "cpu": None,
        "cpl": None,
        "cpk": None,
        "yield": None,
        "sample_ok": 0,
        "sample_ng": 0,
        "status": "N/A",
    }
    if not samples:
        return stats

    mean = statistics.mean(samples)
    stddev = safe_stdev(samples)
    stats.update({
        "max": max(samples),
        "min": min(samples),
        "mean": mean,
        "median": statistics.median(samples),
        "stddev": stddev,
    })

    ng = 0
    for value in samples:
        if (lower is not None and value < lower) or (upper is not None and value > upper):
            ng += 1
    stats["sample_ng"] = ng
    stats["sample_ok"] = len(samples) - ng
    stats["status"] = "NG" if ng else "OK"

    if stddev is None or stddev == 0:
        return stats

    if lower is not None and upper is not None:
        stats["cp"] = (upper - lower) / (6.0 * stddev)
        stats["cpu"] = (upper - mean) / (3.0 * stddev)
        stats["cpl"] = (mean - lower) / (3.0 * stddev)
        stats["cpk"] = min(stats["cpu"], stats["cpl"])
        stats["yield"] = norm_cdf((upper - mean) / stddev) - norm_cdf((lower - mean) / stddev)
    elif upper is not None:
        stats["cpu"] = (upper - mean) / (3.0 * stddev)
        stats["cpk"] = stats["cpu"]
        stats["yield"] = norm_cdf((upper - mean) / stddev)
    elif lower is not None:
        stats["cpl"] = (mean - lower) / (3.0 * stddev)
        stats["cpk"] = stats["cpl"]
        stats["yield"] = 1.0 - norm_cdf((lower - mean) / stddev)
    return stats


def calc_capability_from_stats(mean: Optional[float], stddev: Optional[float], lower: Optional[float], upper: Optional[float], mode: str = "auto") -> Dict[str, Any]:
    capability = {
        "proposed_cp": None,
        "proposed_cpu": None,
        "proposed_cpl": None,
        "proposed_cpk": None,
        "proposed_yield": None,
    }
    if mean is None or stddev in (None, 0):
        return capability

    if mode == "upper" and upper is not None:
        capability["proposed_cpu"] = (upper - mean) / (3.0 * stddev)
        capability["proposed_cpk"] = capability["proposed_cpu"]
        capability["proposed_yield"] = norm_cdf((upper - mean) / stddev)
    elif mode == "lower" and lower is not None:
        capability["proposed_cpl"] = (mean - lower) / (3.0 * stddev)
        capability["proposed_cpk"] = capability["proposed_cpl"]
        capability["proposed_yield"] = 1.0 - norm_cdf((lower - mean) / stddev)
    elif lower is not None and upper is not None:
        capability["proposed_cp"] = (upper - lower) / (6.0 * stddev)
        capability["proposed_cpu"] = (upper - mean) / (3.0 * stddev)
        capability["proposed_cpl"] = (mean - lower) / (3.0 * stddev)
        capability["proposed_cpk"] = min(capability["proposed_cpu"], capability["proposed_cpl"])
        capability["proposed_yield"] = norm_cdf((upper - mean) / stddev) - norm_cdf((lower - mean) / stddev)
    elif upper is not None:
        capability["proposed_cpu"] = (upper - mean) / (3.0 * stddev)
        capability["proposed_cpk"] = capability["proposed_cpu"]
        capability["proposed_yield"] = norm_cdf((upper - mean) / stddev)
    elif lower is not None:
        capability["proposed_cpl"] = (mean - lower) / (3.0 * stddev)
        capability["proposed_cpk"] = capability["proposed_cpl"]
        capability["proposed_yield"] = 1.0 - norm_cdf((lower - mean) / stddev)
    return capability


def tolerance_proposal(record: Dict[str, Any], target_cpk: float) -> Dict[str, Any]:
    mean = record.get("mean")
    stddev = record.get("stddev")
    nominal = record.get("nominal")
    dim = clean_text(record.get("dimension_type")).upper()
    current_plus = record.get("tol_plus")
    current_minus = record.get("tol_minus")
    proposal = {
        "proposed_tol_plus": None,
        "proposed_tol_minus": None,
        "proposed_usl": None,
        "proposed_lsl": None,
        "symmetric_tol": None,
        "proposed_cp": None,
        "proposed_cpu": None,
        "proposed_cpl": None,
        "proposed_cpk": None,
        "proposed_yield": None,
        "proposal_note": "",
    }
    if mean is None or stddev in (None, 0):
        proposal["proposal_note"] = "No proposal: missing samples or stddev=0"
        return proposal

    margin = 3.0 * target_cpk * stddev
    if dim in ("GD&T", "GDT"):
        upper = mean + margin
        proposal.update({
            "proposed_tol_plus": upper,
            "proposed_tol_minus": None,
            "proposed_usl": upper,
            "proposed_lsl": 0.0,
            "proposal_note": "GD&T one-sided upper proposal",
        })
        proposal.update(calc_capability_from_stats(mean, stddev, proposal["proposed_lsl"], proposal["proposed_usl"], mode="upper"))
        return proposal
    if dim == "MAX":
        upper = mean + margin
        proposal.update({
            "proposed_tol_plus": upper,
            "proposed_usl": upper,
            "proposal_note": "MAX one-sided upper proposal",
        })
        proposal.update(calc_capability_from_stats(mean, stddev, proposal["proposed_lsl"], proposal["proposed_usl"], mode="upper"))
        return proposal
    if dim == "MIN":
        lower = mean - margin
        proposal.update({
            "proposed_tol_plus": lower,
            "proposed_lsl": lower,
            "proposal_note": "MIN one-sided lower proposal",
        })
        proposal.update(calc_capability_from_stats(mean, stddev, proposal["proposed_lsl"], proposal["proposed_usl"], mode="lower"))
        return proposal
    if nominal is None:
        proposal["proposal_note"] = "No proposal: missing nominal"
        return proposal

    required_plus = mean + margin - nominal
    required_minus = mean - margin - nominal
    proposed_plus = max(current_plus, required_plus) if current_plus is not None else required_plus
    proposed_minus = min(current_minus, required_minus) if current_minus is not None else required_minus
    symmetric_tol = abs(mean - nominal) + margin
    proposal.update({
        "proposed_tol_plus": proposed_plus,
        "proposed_tol_minus": proposed_minus,
        "proposed_usl": nominal + proposed_plus,
        "proposed_lsl": nominal + proposed_minus,
        "symmetric_tol": symmetric_tol,
        "proposal_note": "Tolerance proposal with nominal fixed",
    })
    proposal.update(calc_capability_from_stats(mean, stddev, proposal["proposed_lsl"], proposal["proposed_usl"]))
    return proposal


def parse_part_info(reader: ExcelXmlReader, preferred_sheets: List[str]) -> Dict[str, Any]:
    info = {
        "part_no": "",
        "part_description": "",
        "supplier": "",
        "revision": "",
        "date": "",
    }
    for sheet in preferred_sheets:
        if sheet not in reader.sheet_targets:
            continue
        cache = {row_num: row for row_num, row in reader.iter_rows(sheet) if row_num <= 6}
        info["part_no"] = clean_text(cache.get(4, {}).get(4))
        info["part_description"] = clean_text(cache.get(5, {}).get(4))
        info["supplier"] = clean_text(cache.get(4, {}).get(14) or cache.get(4, {}).get(8))
        info["revision"] = clean_text(cache.get(6, {}).get(4))
        info["date"] = clean_text(cache.get(6, {}).get(14) or cache.get(6, {}).get(8))
        if any(info.values()):
            return info
    return info


def parse_cpk_workbook(path: Path, target_cpk: float) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    with ExcelXmlReader(path) as reader:
        part_info = parse_part_info(reader, [s for s in reader.sheet_names() if s.startswith("CPK CAV")] + reader.sheet_names())
        sheet_names = [s for s in reader.sheet_names() if s.startswith("CPK CAV") or s.startswith("Raw data")]
        for sheet_name in sheet_names:
            kind = "CPK" if sheet_name.startswith("CPK CAV") else "Raw data"
            cavity = normalize_cavity(sheet_name)
            for row_num, row in reader.iter_rows(sheet_name):
                if row_num < 11:
                    continue
                fai_no = clean_text(row.get(3))
                if not fai_no:
                    continue
                spc_no = clean_text(row.get(2))
                dim_type = clean_text(row.get(6))
                nominal = as_float(row.get(7))
                tol_plus = as_float(row.get(8))
                tol_minus = as_float(row.get(9))
                lower, upper = compute_limits(dim_type, nominal, tol_plus, tol_minus)
                samples = []
                for col in range(37, 211):  # AK:HB; template supports many samples, stop after first long blank tail is not required.
                    value = as_float(row.get(col))
                    if value is not None:
                        samples.append(value)
                stats = calc_stats(samples, lower, upper)
                record: Dict[str, Any] = {
                    "source_file": path.name,
                    "file_tag": file_tag_text(path.name),
                    "report_type": "CPK",
                    "sheet_kind": kind,
                    "sheet_name": sheet_name,
                    "row": row_num,
                    "cavity": cavity,
                    "part_no": part_info.get("part_no", ""),
                    "part_description": part_info.get("part_description", ""),
                    "revision": part_info.get("revision", ""),
                    "spc_no": spc_no,
                    "fai_no": fai_no,
                    "description": clean_text(row.get(4)),
                    "target_cpk": as_float(row.get(5)) or target_cpk,
                    "dimension_type": dim_type,
                    "nominal": nominal,
                    "tol_plus": tol_plus,
                    "tol_minus": tol_minus,
                    "usl": upper,
                    "lsl": lower,
                    "inspection_method": clean_text(row.get(10)),
                    "gauged_100pct": clean_text(row.get(11)),
                    "samples": samples,
                    "samples_json": json.dumps(samples, ensure_ascii=False),
                }
                record.update(stats)
                record.update(tolerance_proposal(record, target_cpk))
                record["cpk_status"] = "N/A" if record.get("cpk") is None else ("NG" if record["cpk"] < target_cpk else "OK")
                records.append(record)
    metadata = {"file": path.name, "type": "CPK", "records": len(records)}
    return records, metadata


def parse_fai_workbook(path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    with ExcelXmlReader(path) as reader:
        fai_sheets = [s for s in reader.sheet_names() if re.match(r"N[A-Za-z]\s*\d+", s, flags=re.I)]
        part_info = parse_part_info(reader, fai_sheets + reader.sheet_names())
        for sheet_name in fai_sheets:
            cavity = normalize_cavity(sheet_name)
            for row_num, row in reader.iter_rows(sheet_name):
                if row_num < 11:
                    continue
                fai_no = clean_text(row.get(2))
                if not fai_no:
                    continue
                spc_no = clean_text(row.get(3))
                dim_type = clean_text(row.get(4))
                nominal = as_float(row.get(5))
                tol_plus = as_float(row.get(6))
                tol_minus = as_float(row.get(7))
                lower, upper = compute_limits(dim_type, nominal, tol_plus, tol_minus)
                samples = []
                for col in range(11, 14):  # K:M in the provided FAI template.
                    value = as_float(row.get(col))
                    if value is not None:
                        samples.append(value)
                stats = calc_stats(samples, lower, upper)
                if stats["sample_count"] != 3 or stats["sample_ng"]:
                    stats["status"] = "NG"
                record: Dict[str, Any] = {
                    "source_file": path.name,
                    "file_tag": file_tag_text(path.name),
                    "report_type": "FAI",
                    "sheet_kind": "FAI",
                    "sheet_name": sheet_name,
                    "row": row_num,
                    "cavity": cavity,
                    "part_no": part_info.get("part_no", ""),
                    "part_description": part_info.get("part_description", ""),
                    "revision": part_info.get("revision", ""),
                    "spc_no": spc_no,
                    "fai_no": fai_no,
                    "description": "",
                    "target_cpk": None,
                    "dimension_type": dim_type,
                    "nominal": nominal,
                    "tol_plus": tol_plus,
                    "tol_minus": tol_minus,
                    "usl": upper,
                    "lsl": lower,
                    "inspection_method": clean_text(row.get(8)),
                    "gauged_100pct": "",
                    "samples": samples,
                    "samples_json": json.dumps(samples, ensure_ascii=False),
                }
                record.update(stats)
                record["item_status"] = record["status"]
                records.append(record)
    metadata = {"file": path.name, "type": "FAI", "records": len(records)}
    return records, metadata


def find_input_file_lists(input_dir: Path, cpk_file: Optional[str], fai_file: Optional[str], require_cpk: bool = True, require_fai: bool = True) -> Tuple[List[Path], List[Path]]:
    if cpk_file:
        cpk_path = Path(cpk_file)
        if not cpk_path.is_absolute():
            cpk_path = input_dir / cpk_path
        cpk_paths = [cpk_path]
    else:
        cpk_paths = sorted(p for p in input_dir.glob("*.xls*") if "CPK" in p.name.upper() and not p.name.startswith(".~") and not p.name.startswith("~$"))
        if require_cpk and not cpk_paths:
            raise FileNotFoundError("No CPK Excel file found. Use --cpk-file.")

    if fai_file:
        fai_path = Path(fai_file)
        if not fai_path.is_absolute():
            fai_path = input_dir / fai_path
        fai_paths = [fai_path]
    else:
        fai_paths = sorted(p for p in input_dir.glob("*.xls*") if "FAI" in p.name.upper() and not p.name.startswith(".~") and not p.name.startswith("~$"))
        if require_fai and not fai_paths:
            raise FileNotFoundError("No FAI Excel file found. Use --fai-file.")

    for path in cpk_paths + fai_paths:
        if not path.exists():
            raise FileNotFoundError(str(path))
    return cpk_paths, fai_paths


def find_input_files(input_dir: Path, cpk_file: Optional[str], fai_file: Optional[str]) -> Tuple[Path, Path]:
    cpk_paths, fai_paths = find_input_file_lists(input_dir, cpk_file, fai_file)
    return cpk_paths[0], fai_paths[0]


def parse_cpk_workbooks(paths: List[Path], target_cpk: float) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    files = []
    for path in paths:
        file_records, _ = parse_cpk_workbook(path, target_cpk)
        records.extend(file_records)
        files.append(path.name)
    return records, {"files": files, "type": "CPK", "records": len(records)}


def parse_fai_workbooks(paths: List[Path]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    files = []
    for path in paths:
        file_records, _ = parse_fai_workbook(path)
        records.extend(file_records)
        files.append(path.name)
    return records, {"files": files, "type": "FAI", "records": len(records)}


def round_float(value: Any, digits: int = 6) -> Any:
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return round(value, digits)
    return value


def add_sheet(wb: Workbook, title: str, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            ws.cell(row_idx, col_idx, round_float(value))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = defaultdict(lambda: 10)
    for row in ws.iter_rows():
        for cell in row:
            text = clean_text(cell.value)
            widths[cell.column] = min(max(widths[cell.column], len(text) + 2), 60)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def source_label(source: Any) -> str:
    if isinstance(source, (list, tuple)):
        names = [Path(item).name for item in source]
        if not names:
            return ""
        if len(names) <= 5:
            return "; ".join(names)
        return "%s files: %s ..." % (len(names), "; ".join(names[:5]))
    return Path(source).name


def group_cpk_spc_status(cpk_records: List[Dict[str, Any]], target_cpk: float) -> List[Dict[str, Any]]:
    groups: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for record in cpk_records:
        if record.get("sheet_kind") != "CPK":
            continue
        spc_no = clean_text(record.get("spc_no"))
        fai_no = clean_text(record.get("fai_no"))
        if not spc_no or not fai_no:
            continue
        key = (clean_text(record.get("source_file")), spc_no, fai_no)
        groups[key].append(record)
    rows = []
    for (source_file, spc_no, fai_no), records in sorted(groups.items()):
        cpk_values = [r.get("cpk") for r in records if r.get("cpk") is not None]
        is_ng = any(value < target_cpk for value in cpk_values) or not cpk_values
        rows.append({
            "source_file": source_file,
            "spc_no": spc_no,
            "fai_no": fai_no,
            "row_count": len(records),
            "min_cpk": min(cpk_values, default=None),
            "spc_status": "NG" if is_ng else "OK",
        })
    return rows


def build_summary_rows(cpk_records: List[Dict[str, Any]], fai_records: List[Dict[str, Any]], target_cpk: float, cpk_path: Any, fai_path: Any) -> List[Dict[str, Any]]:
    fai_item_ok = sum(1 for r in fai_records if r.get("status") == "OK")
    fai_item_ng = sum(1 for r in fai_records if r.get("status") == "NG")
    fai_sample_ok = sum(int(r.get("sample_ok") or 0) for r in fai_records)
    fai_sample_ng = sum(int(r.get("sample_ng") or 0) for r in fai_records)
    cpk_table_records = [r for r in cpk_records if r.get("sheet_kind") == "CPK"]
    cpk_with_value = [r for r in cpk_table_records if r.get("cpk") is not None]
    low_cpk = [r for r in cpk_with_value if r["cpk"] < target_cpk]
    spc_status = group_cpk_spc_status(cpk_records, target_cpk)
    rows = [
        {"Metric": "Generated At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Metric": "CPK File", "Value": source_label(cpk_path)},
        {"Metric": "FAI File", "Value": source_label(fai_path)},
        {"Metric": "Target CPK", "Value": target_cpk},
        {"Metric": "CPK Dimension Rows", "Value": len(cpk_table_records)},
        {"Metric": "CPK Rows With CPK", "Value": len(cpk_with_value)},
        {"Metric": "CPK < Target Rows", "Value": len(low_cpk)},
        {"Metric": "Minimum CPK", "Value": min((r["cpk"] for r in cpk_with_value), default=None)},
        {"Metric": "FAI Dimension Rows", "Value": len(fai_records)},
        {"Metric": "FAI Item OK", "Value": fai_item_ok},
        {"Metric": "FAI Item NG", "Value": fai_item_ng},
        {"Metric": "FAI Sample OK", "Value": fai_sample_ok},
        {"Metric": "FAI Sample NG", "Value": fai_sample_ng},
        {"Metric": "SPC OK", "Value": sum(1 for r in spc_status if r["spc_status"] == "OK")},
        {"Metric": "SPC NG", "Value": sum(1 for r in spc_status if r["spc_status"] == "NG")},
    ]
    return rows


def group_spc_status(fai_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for record in fai_records:
        groups[(clean_text(record.get("cavity")), clean_text(record.get("spc_no")))].append(record)
    rows = []
    for (cavity, spc_no), records in sorted(groups.items()):
        ng_items = [r for r in records if r.get("status") == "NG"]
        rows.append({
            "cavity": cavity,
            "spc_no": spc_no,
            "fai_count": len({r.get("fai_no") for r in records}),
            "item_count": len(records),
            "sample_count": sum(int(r.get("sample_count") or 0) for r in records),
            "sample_ng": sum(int(r.get("sample_ng") or 0) for r in records),
            "spc_status": "NG" if ng_items else "OK",
            "ng_fai_list": ", ".join(sorted({clean_text(r.get("fai_no")) for r in ng_items})),
        })
    return rows


def worst_cavity_rows(cpk_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for record in cpk_records:
        if record.get("cpk") is None:
            continue
        spc_no = clean_text(record.get("spc_no"))
        if not spc_no:
            continue
        groups[spc_no].append(record)
    rows = []
    for spc_no, records in sorted(groups.items()):
        worst = min(records, key=lambda r: r.get("cpk"))
        rows.append({
            "spc_no": spc_no,
            "worst_sheet_kind": worst.get("sheet_kind"),
            "worst_fai_no": worst.get("fai_no"),
            "description": worst.get("description"),
            "worst_cavity": worst.get("cavity"),
            "worst_cpk": worst.get("cpk"),
            "worst_mean": worst.get("mean"),
            "worst_stddev": worst.get("stddev"),
            "sample_count": worst.get("sample_count"),
        })
    return rows


def export_workbook(output_path: Path, cpk_records: List[Dict[str, Any]], fai_records: List[Dict[str, Any]], target_cpk: float, cpk_path: Any, fai_path: Any) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    summary_rows = build_summary_rows(cpk_records, fai_records, target_cpk, cpk_path, fai_path)
    add_sheet(wb, "Summary", ["Metric", "Value"], summary_rows)

    cpk_headers = [
        "source_file", "file_tag", "sheet_kind", "sheet_name", "cavity", "row", "part_no", "part_description", "revision",
        "spc_no", "fai_no", "description", "dimension_type", "nominal", "tol_plus", "tol_minus", "usl", "lsl",
        "inspection_method", "sample_count", "min", "max", "mean", "median", "stddev", "cp", "cpu", "cpl", "cpk",
        "yield", "sample_ok", "sample_ng", "status", "cpk_status",
    ]
    add_sheet(wb, "CPK_Summary", cpk_headers, cpk_records)

    low_records = [r for r in cpk_records if r.get("cpk") is not None and r.get("cpk") < target_cpk]
    add_sheet(wb, "CPK_LowRisk", cpk_headers, low_records)

    proposal_headers = [
        "source_file", "file_tag", "sheet_kind", "cavity", "spc_no", "fai_no", "description", "dimension_type", "nominal", "tol_plus", "tol_minus",
        "mean", "stddev", "cpk", "proposed_tol_plus", "proposed_tol_minus", "proposed_usl", "proposed_lsl",
        "proposed_cp", "proposed_cpu", "proposed_cpl", "proposed_cpk", "proposed_yield", "symmetric_tol", "proposal_note",
    ]
    add_sheet(wb, "Tolerance_Proposal", proposal_headers, low_records)

    fai_headers = [
        "source_file", "file_tag", "sheet_name", "cavity", "row", "part_no", "part_description", "revision", "spc_no", "fai_no",
        "dimension_type", "nominal", "tol_plus", "tol_minus", "usl", "lsl", "inspection_method", "sample_count",
        "min", "max", "mean", "median", "stddev", "sample_ok", "sample_ng", "status", "samples_json",
    ]
    add_sheet(wb, "FAI_OK_NG", fai_headers, fai_records)

    spc_rows = group_spc_status(fai_records)
    add_sheet(wb, "SPC_OK_NG", ["cavity", "spc_no", "fai_count", "item_count", "sample_count", "sample_ng", "spc_status", "ng_fai_list"], spc_rows)

    cpk_spc_rows = group_cpk_spc_status(cpk_records, target_cpk)
    add_sheet(wb, "CPK_SPC_OK_NG", ["source_file", "spc_no", "fai_no", "row_count", "min_cpk", "spc_status"], cpk_spc_rows)

    worst_rows = worst_cavity_rows(cpk_records)
    add_sheet(wb, "Worst_Cavity", ["spc_no", "worst_sheet_kind", "worst_fai_no", "description", "worst_cavity", "worst_cpk", "worst_mean", "worst_stddev", "sample_count"], worst_rows)

    raw_headers = cpk_headers + ["samples_json"]
    add_sheet(wb, "Raw_Normalized_CPK", raw_headers, cpk_records)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="PQE Phase 1 MVP: CPK/FAI parser and summary exporter")
    parser.add_argument("--input-dir", default=".", help="Directory containing the CPK and FAI Excel files")
    parser.add_argument("--cpk-file", default=None, help="CPK report file name/path. Auto-detected if omitted.")
    parser.add_argument("--fai-file", default=None, help="FAI report file name/path. Auto-detected if omitted.")
    parser.add_argument("--output", default=None, help="Output xlsx path")
    parser.add_argument("--target-cpk", type=float, default=1.33, help="CPK threshold, default 1.33")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).resolve()
    cpk_paths, fai_paths = find_input_file_lists(input_dir, args.cpk_file, args.fai_file, require_cpk=False, require_fai=False)
    if not cpk_paths and not fai_paths:
        raise FileNotFoundError("No CPK or FAI Excel file found.")
    if args.output:
        output_path = Path(args.output)
        if not output_path.is_absolute():
            output_path = input_dir / output_path
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = input_dir / ("PQE_Phase1_Summary_%s.xlsx" % stamp)

    try:
        cpk_records, cpk_meta = parse_cpk_workbooks(cpk_paths, args.target_cpk)
        fai_records, fai_meta = parse_fai_workbooks(fai_paths)
        export_workbook(output_path, cpk_records, fai_records, args.target_cpk, cpk_paths, fai_paths)
    except BadZipFile as exc:
        raise SystemExit("Invalid Excel file: %s" % exc)

    low_count = sum(1 for r in cpk_records if r.get("cpk") is not None and r.get("cpk") < args.target_cpk)
    print("CPK records: %s" % cpk_meta["records"])
    print("FAI records: %s" % fai_meta["records"])
    print("CPK files: %s" % len(cpk_paths))
    print("FAI files: %s" % len(fai_paths))
    print("CPK < %.3f: %s" % (args.target_cpk, low_count))
    print("Output: %s" % output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())